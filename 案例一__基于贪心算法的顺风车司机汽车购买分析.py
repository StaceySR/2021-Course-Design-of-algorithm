import requests
import time
import xlrd
import openpyxl
from openpyxl import load_workbook
from bs4 import BeautifulSoup as bs


def common_crawler(url):
    '''
    通用的爬取网页源代码的函数
    :param url: 网页链接
    :return: 网页源代码
    '''
    header = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit'
        '/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36'
    }
    while True:  # 一直循环，直到访问站点成功
        try:
            # 以下except都是用来捕获当requests请求出现异常时，
            # 通过捕获然后等待网络情况的变化，以此来保护程序的不间断运行
            resp = requests.get(url, headers=header, timeout=20)
            break
        except requests.exceptions.ConnectionError:  # 网络连接异常
            print('ConnectionError -- please wait 3 seconds')
            time.sleep(3)
        except requests.exceptions.ChunkedEncodingError:
            print('ChunkedEncodingError -- please wait 3 seconds')
            time.sleep(3)
        except:
            print('Unfortunitely -- An Unknow Error Happened, Please wait 3 seconds')
            time.sleep(3)
            time.sleep(3)

    html_data = resp.text.encode(resp.encoding)
    # 使用BeautifulCoup库进行HTML代码的解析。
    # 第一个参数为需要提取数据的html，第二个参数是指定解析器
    soup = bs(html_data, 'html.parser')
    return soup


def find_right_car(low, high, seat_num):
    '''
    :param low: 司机能承受的汽车价格范围的最低价
    :param high: 司机能承受的汽车价格范围的最高价
    :param seat_num: 汽车座位数
    '''
    # 汽车座位一般只分到8座及以上
    if seat_num >= 7:
        seat_num = 7
    url = 'http://price.cheshi.com/select/' + str(low) + '_' + str(high) + '-0-0-0-0-0-0-0-0-0-' + str(seat_num) +'-1-0'
    soup = common_crawler(url)
    page = soup.find('div', class_='page')
    hrefs = page.find_all('a', target='_self')
    if len(hrefs) > 1:
        for href in hrefs:
            page_url = 'http://price.cheshi.com' + href['href']
            print_all_car(page_url)
    else:
        # 爬取汽车信息
        print_all_car(url)


def print_all_car(url):
    '''
    寻找所有满足司机要求的汽车信息的函数
    :param url: 汽车网页链接
    :return: 无，直接输出满足要求的汽车信息
    '''
    soup = common_crawler(url)
    list = soup.find('div', class_='list')
    # http://price.cheshi.com/series_129.html  战旗
    if list.find_all('dl'):
        dls = list.find_all('dl')
        index = 1
        for dl in dls:
            car_name = dl.find('h3').text
            car_url = ' http://price.cheshi.com' + dl.find('h3').find('a')['href']
            car_price = dl.find_all('p')[1].text
            print(index, car_name, car_url, car_price)
            output_CarData(index, car_name, car_url, car_price)
            index += 1
    else:  # 没有满足条件的汽车信息
        print('None')


def carPooling(trips, capacity, map):
    '''
    :param trips: 顺风车行程订单
    :param capacity: 司机汽车座位数
    :param map: 司机回家路线图
    :return: 如果车座数满足接到全部顺风车订单的要求，就返回车座数，否则返回False
    '''
    travel = {}
    for i in trips:
        if i[1] in travel:
            travel[map[i[1]]] += i[0]
        else:
            travel[map[i[1]]] = i[0]
        if i[2] in travel:
            travel[map[i[2]]] -= i[0]
        else:
            travel[map[i[2]]] = -i[0]
    l = sorted(travel)
    people_in_car = 0
    for i in l:
        people_in_car += travel[i]
        if people_in_car > capacity:
            # print('根据给出的行程计划表和车子的座位数，在初步排除掉不顺路的订单，剩下的订单不能全部完成')
            return False
    print('根据给出的行程计划表和车子的座位数{}，在初步排除掉不顺路的订单，剩下的订单可以全部完成。'.format(capacity))
    return True


# 输入数据
# 1、读取路线图map信息
def load_mapData():
    workbook = xlrd.open_workbook(r'案例一Data.xlsx')
    table = workbook.sheet_by_name('路线')
    rows = table.nrows  # 获取工作表的行数
    map = {}
    for row in range(1, rows):
        map[table.cell(row, 0).value] = int(table.cell(row, 1).value)
    return map


# 2、读取顺风车形成订单信息trips
def load_tripsData():
    workbook = xlrd.open_workbook(r'案例一Data.xlsx')
    table = workbook.sheet_by_name('订单')
    rows = table.nrows  # 获取工作表的行数
    trips = []
    for row in range(1, rows):
        passenage_num = int(table.cell(row, 0).value)
        get_on_station = table.cell(row, 1).value
        get_off_station = table.cell(row, 2).value
        one_trip = [passenage_num, get_on_station, get_off_station]
        trips.append(one_trip)
    return trips


# 3、读取司机可承受的汽车价格范围
def load_priceData():
    workbook = xlrd.open_workbook(r'案例一Data.xlsx')
    table = workbook.sheet_by_name('司机能承受汽车价格范围')
    low = table.cell(1, 0).value
    high = table.cell(1, 1).value
    return low, high


# 输出数据-------符合要求的汽车信息
def output_CarData(index, car_name, car_url, car_price):
    workbook = load_workbook("案例一Data.xlsx")  # 生成一个已存在的workbook对象
    wk_name = workbook.sheetnames
    wk_sheet = workbook[wk_name[-1]]
    wk_sheet = workbook.active  # 激活sheet
    wk_sheet.cell(index+2, 1).value = car_name
    wk_sheet.cell(index+2, 2).value = car_url
    wk_sheet.cell(index+2, 3).value = car_price[4:]
    workbook.save("案例一Data.xlsx")  # 保存


if __name__ == '__main__':

    # 司机下班回家路线  地名+处在路线中的顺序
    map = load_mapData()

    # 顺风车行程订单列表
    #        乘客数  上车地    下车地
    trips = load_tripsData()
    # 司机可承受汽车价格范围 [最低价low,最高价high]
    low, high = load_priceData()
    flag = False
    for i in range(2, 8):
        capacity = i
        result = carPooling(trips, capacity, map)
        if result:
            print('\n--------------以下汽车满足司机要求，可选---------------')
            find_right_car(low, high, capacity)
            flag = True
            break
    if not flag:
        find_right_car(low, high, 7)  # 如果座位数为7座还是无法接到所有顺风车订单，那么就选最大的家庭轿车最大的座位数7


